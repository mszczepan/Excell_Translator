import openpyxl


def open_file(path):
    wb = openpyxl.load_workbook(path)
    return wb


def open_sheet(wb, sheet):
    ws = wb[sheet]
    return ws


def read_file(ws, column, start_row, end_row):
    read_list = [ws[column + str(row)].value for row in range(start_row, end_row)]
    return read_list


def write_file(ws, column_nr, column_sym, start_row, read_list):
    ws.insert_cols(column_nr)
    row_nr = start_row
    for row_value in read_list:
        ws[column_sym + row_nr] = row_value
        row_nr = row_nr + 1



